import os
import json
import codecs
from openpyxl import load_workbook
import pandas as pd
from datetime import datetime

# ====================================================================================
# 1. CONFIGURACIÓN DE RUTAS LOCALES EXACTAS
# ====================================================================================
RUTA_CARPETA = r"C:\Users\Coord Planeación\Documents\DASHBOARD VIGENTES"
ARCHIVO_EXCEL = os.path.join(RUTA_CARPETA, "vigentes dashboard.xlsx")
ARCHIVO_JS_DATOS = os.path.join(RUTA_CARPETA, "datos.js")

print("📖 Abriendo el Excel masivo de 65 MB en modo Ultra-Light (Read Only)...")

# Diccionario maestro para consolidar las agrupaciones de 6 dimensiones y sumas
matriz_consolidada = {}

try:
    # Mantenemos read_only=True para proteger tu memoria RAM
    wb = load_workbook(filename=ARCHIVO_EXCEL, read_only=True, data_only=True)
    ws = wb['DATA'] if 'DATA' in wb.sheetnames else wb.active
except Exception as e:
    print(f"❌ Error al abrir el archivo Excel. Detalle: {e}")
    input("\nPresiona Enter para salir...")
    exit()

print("🧼 Analizando encabezados y mapeando las 6 columnas...")
iterador_filas = ws.iter_rows(values_only=True)
encabezados = next(iterador_filas)

# Limpieza estricta de encabezados para evitar problemas con tildes o espacios
encabezados_limpios = [str(h).strip().upper().replace('🗠', '').replace('É', 'E').replace('Ó', 'O') for h in encabezados]

def buscar_indice_columna(nombres_posibles, indice_defecto):
    for nombre in nombres_posibles:
        if nombre in encabezados_limpios:
            return encabezados_limpios.index(nombre)
    for i, h in enumerate(encabezados_limpios):
        if any(p in h for p in nombres_posibles):
            return i
    return indice_defecto

# Mapeo automático de las 6 dimensiones + Columna P (Cuota)
idx_estatus = buscar_indice_columna(['ESTATUS', 'ESTADO'], 0)
idx_tipo = buscar_indice_columna(['TIPO', 'CANAL', 'COBRADOR'], 1)
idx_regional = buscar_indice_columna(['REGIONAL', 'REGION', 'ZONA'], 2)
idx_atraso = buscar_indice_columna(['GRUPO ATRASO', 'ATRASO', 'RANGO'], 3)
idx_gestion = buscar_indice_columna(['GESTION', 'GESTIONANDO', 'ESTADO GESTION'], 4) 
idx_tipovendedor = buscar_indice_columna(['TIPOVENDEDOR', 'TIPO VENDEDOR', 'VENDEDOR'], -1)
idx_producto = buscar_indice_columna(['PRODUCTO', 'PROD'], -1)
idx_zona = buscar_indice_columna(['ZONAS', 'ZONA', 'MUNICIPIO', 'CIUDAD'], -1)
idx_nombres = buscar_indice_columna(['NOMBRES', 'NOMBRE', 'CLIENTE'], 5)
idx_direccion = buscar_indice_columna(['DIRECCION', 'DIR'], 10)
idx_cuota = buscar_indice_columna(['CUOTA', 'MONTO', 'VALOR'], 15) # Columna P física por defecto

print(f"🔍 Coordenadas físicas enlazadas con éxito:")
print(f"   - Estatus:        Índice {idx_estatus} ({encabezados[idx_estatus]})")
print(f"   - Tipo:           {idx_tipo} ({encabezados[idx_tipo]})")
print(f"   - Regional:       {idx_regional} ({encabezados[idx_regional]})")
print(f"   - Atraso:         {idx_atraso} ({encabezados[idx_atraso]})")
print(f"   - Gestión:        {idx_gestion} ({encabezados[idx_gestion]})") 
if idx_tipovendedor != -1: print(f"   - Tipo Vendedor:  {idx_tipovendedor} ({encabezados[idx_tipovendedor]})")
if idx_producto != -1: print(f"   - Producto:       {idx_producto} ({encabezados[idx_producto]})")
if idx_zona != -1: print(f"   - Zona:           {idx_zona} ({encabezados[idx_zona]})")
print(f"   - Cuota:          {idx_cuota} ({encabezados[idx_cuota]})")

print("\n🧮 Procesando registros fila por fila en tiempo récord...")
contador_filas = 0

for fila in iterador_filas:
    # Si la fila está vacía o es de relleno al final, la ignoramos
    if fila[idx_estatus] is None and fila[idx_cuota] is None:
        continue
        
    contador_filas += 1
    if contador_filas % 25000 == 0:
        print(f"   ⏳ Registros consolidados: {contador_filas:,}...")

    # Extracción y limpieza segura de textos individuales
    val_estatus = str(fila[idx_estatus]).strip() if fila[idx_estatus] is not None else "N/A"
    val_tipo = str(fila[idx_tipo]).strip() if fila[idx_tipo] is not None else "N/A"
    val_regional = str(fila[idx_regional]).strip() if fila[idx_regional] is not None else "N/A"
    val_atraso = str(fila[idx_atraso]).strip().upper() if fila[idx_atraso] is not None else "N/A"
    val_gestion = str(fila[idx_gestion]).strip() if fila[idx_gestion] is not None else "SIN GESTION"
    val_tipovendedor = str(fila[idx_tipovendedor]).strip() if idx_tipovendedor != -1 and fila[idx_tipovendedor] is not None else "SIN ESPECIFICAR"
    val_producto = str(fila[idx_producto]).strip() if idx_producto != -1 and fila[idx_producto] is not None else "SIN ESPECIFICAR"
    val_zona = str(fila[idx_zona]).strip() if idx_zona != -1 and fila[idx_zona] is not None else "SIN ESPECIFICAR"

    # Limpieza matemática de la cuota
    try:
        monto_limpio = float(fila[idx_cuota])
        if pd.isna(monto_limpio):
            monto_limpio = 0.0
    except:
        monto_limpio = 0.0

    # Llave maestra expandida a 8 dimensiones para agrupar en caliente
    llave_grupo = (val_estatus, val_tipo, val_regional, val_atraso, val_gestion, val_tipovendedor, val_producto, val_zona)
    
    if llave_grupo not in matriz_consolidada:
        matriz_consolidada[llave_grupo] = {"cantidad": 0, "monto": 0.0}
        
    matriz_consolidada[llave_grupo]["cantidad"] += 1
    matriz_consolidada[llave_grupo]["monto"] += monto_limpio



wb.close()

# Estructuración final de la matriz para JavaScript
js_records_final = []
for llave, totales in matriz_consolidada.items():
    js_records_final.append({
        "estatus": llave[0],
        "tipo": llave[1],
        "regional": llave[2],
        "mes_pago": "2026-06",
        "atraso": llave[3],
        "gestion": llave[4], 
        "TIPOVENDEDOR": llave[5],
        "PRODUCTO": llave[6],
        "ZONA": llave[7],
        "cantidad": totales["cantidad"],
        "monto": totales["monto"]
    })

# ====================================================================================
# 3. EXPORTACIÓN DIRECTA COMO VARIABLE JAVASCRIPT CORREGIDA
# ====================================================================================
try:
    timestamp_mod = os.path.getmtime(ARCHIVO_EXCEL)
    fecha_modificacion = datetime.fromtimestamp(timestamp_mod).strftime('%d/%m/%Y %I:%M %p')
except Exception:
    fecha_modificacion = "Desconocida"

try:
    with codecs.open(ARCHIVO_JS_DATOS, "w", encoding="utf-8") as f:
        f.write(f"const FECHA_ACTUALIZACION = '{fecha_modificacion}';\n")
        f.write("const SERVER_DATA = ")
        json.dump(js_records_final, f, ensure_ascii=False, indent=4)
        f.write(";")
    print(f"\n✨ ¡Éxito absoluto! Archivo 'datos.js' actualizado con {contador_filas:,} filas consolidadas.")
except Exception as e:
    print(f"❌ Error al escribir el archivo de datos. Detalle: {e}")

input("\n🚀 Sincronización completada con éxito. Presiona Enter para cerrar...")